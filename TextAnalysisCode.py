import openpyxl
import requests as r
import ssl
from bs4 import BeautifulSoup
from nltk.tokenize import word_tokenize
import pandas as pd
import re

def TextAnalysis(fname):
    words = list()                              #initializing list for words in paragraph body
    stopwords = list()                          #initializing list for words in StopWords_Generic.txt file
    p_rows = list()                             #initializing list for words having non-zero positive dictionary values
    n_rows = list()                             #initializing list for words having non-zero negative dictionary values
    c_rows = list()                             #initializing list for words having complex dictionary words
    lines = list()                              #initializing list for sentences in paragraph body
    poswords = list()                           #initializing list for words in paragraph to check for positive dictionary values
    negwords = list()                           #initializing list for words in paragraph to check for negative dictionary values
    sentences = list()

    personal_pronoun_list = ["i","I","we","We","WE","my","My","MY","ours","Ours","OURS","us","Us","our","Our","OUR","me","Me"]      	 #creating a list of personal pronouns to check and match with regex statements
    data = pd.read_excel('LoughranMcDonald_MasterDictionary_2020.xlsx', header=0)                                                        #reading all the words mentioned in the LoughranMcDonald_MasterDictionary_2020.xlsx file for analysis
    thandle = open("StopWords_Generic.txt","r")                                                                                          #reading all the stopwords mentioned in the StopWords_Generic.txt file

    pos = 0
    neg = 0
    count_prp = 0
    tot_chars = 0

    fhandle = open(fname,"r")

    for line in fhandle:
        a = line.strip()
        pos = a.find("-")
        result = a[pos+1:]
        res = result.strip()
        sentences = res.split(".")                                       #splitting the text into sentences for Readability analysis
        all_words = re.sub('[^a-zA-Z0-9 \n\.]', '', res)                 #taking care of punctuations and extra characters while parsing paragraph data
        all_words = re.sub('[.]', '', all_words)                         #taking care of punctuations and extra characters while parsing paragraph data
        words = word_tokenize(all_words)

    for sentence in sentences:
        s = re.sub(r"^\s+","",sentence,flags=re.I)
        lines.append(s)

    if(lines[len(lines)-1]==""):                                         #Handling conditions for finding average sentence length
        average_sentence_length = len(words) / (len(lines) - 1)
    else:
        average_sentence_length = len(words) / len(lines)

    word_df = pd.DataFrame(data[data['Word'].isin(each_word.upper() for each_word in words)])               #putting all words parsed from html file and checking the dictionary for each word's presence, and extracting their corresponding data into new dataframe

    complex_df = pd.DataFrame(word_df[word_df['Complexity'] != 0])

    c_rows = complex_df.iloc[:, 0].tolist()

    no_complex = len(c_rows)                                               #Number of complex words extracted from text file

    prc_complex_words = no_complex/len(words)                              #Percentage of complex words in the text file

    fog_index = 0.4*(average_sentence_length + prc_complex_words)          #Fog index data extracted from text file

    for word in thandle:                                                   #Creating list of stopwords to perform text cleaning and sentimental analysis
        stop = word.strip()
        stopwords.append(stop)

    words_clean = [w for w in words if w.upper() not in stopwords]          #performing the cleaning operation on the words present in paragraph which are not present in StopWords dictionary

    word_df1 = pd.DataFrame(data[data['Word'].isin(each_word.upper() for each_word in words_clean)])  #putting all words parsed from html file and checking the dictionary for each word's presence, and extracting their corresponding data into new dataframe

    positive = pd.DataFrame(word_df1[word_df1['Positive'] != 0])
    negative = pd.DataFrame(word_df1[word_df1['Negative'] != 0])

    p_rows = positive.iloc[:, 0].tolist()     #Extracting list of words with positive values != 0
    n_rows = negative.iloc[:, 0].tolist()     #Extracting list of words with negative values != 0

    for word in words_clean:
        each_word = word.upper()
        if each_word in p_rows:
            poswords.append(each_word)
            pos = pos + 1
        elif each_word in n_rows:
            negwords.append(each_word)
            neg = neg - 1
        else:
            continue

    neg = len(negwords)
    pos = len(poswords)

    polarity = (pos - neg)/((pos + neg) + 0.000001)

    subjectivity = (pos + neg)/(len(words_clean)+0.000001)

    tot_syll = word_df1["Syllables"].sum()

    syll_count_per_word = tot_syll/len(words_clean)

    for word in words:
        for pers_pronoun in personal_pronoun_list:
            if re.match(word,pers_pronoun):
                count_prp = count_prp + 1
            else:
                continue

    for word in words:
        chars = len(word)
        tot_chars = tot_chars + chars

    average_word_length = tot_chars/len(words)

    output_df = pd.DataFrame(columns = ['URL_ID','URL','POSITIVE SCORE','NEGATIVE SCORE','POLARITY SCORE','SUBJECTIVITY SCORE','AVG SENTENCE LENGTH','PERCENTAGE OF COMPLEX WORDS (in %)','FOG INDEX','AVG NUMBER OF WORDS PER SENTENCE','COMPLEX WORD COUNT','WORD COUNT','SYLLABLE PER WORD','PERSONAL PRONOUNS','AVG WORD LENGTH'])

    output_df = output_df.append({'URL_ID': name,'URL': link, 'POSITIVE SCORE' : pos, 'NEGATIVE SCORE' : neg, 'POLARITY SCORE' : polarity, 'SUBJECTIVITY SCORE': subjectivity, 'AVG SENTENCE LENGTH': average_sentence_length, 'PERCENTAGE OF COMPLEX WORDS (in %)': (prc_complex_words*100), 'FOG INDEX': fog_index, 'AVG NUMBER OF WORDS PER SENTENCE': average_sentence_length, 'COMPLEX WORD COUNT': no_complex, 'WORD COUNT': len(words_clean), 'SYLLABLE PER WORD': syll_count_per_word, 'PERSONAL PRONOUNS': count_prp, 'AVG WORD LENGTH': average_word_length},ignore_index = True)

    output_df.to_csv('OutputData.csv', mode = 'a', index = False, header = None)



headers = {
    'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10.12; rv:55.0) Gecko/20100101 Firefox/55.0',  #adding headers to specify user access to HTTPS site
}


# Ignore SSL certificate errors
ctx = ssl.create_default_context()
ctx.check_hostname = False
ctx.verify_mode = ssl.CERT_NONE

links = list()                               #initializing list for URLs
names = list()                               #initializing list for URL_IDs

wb = openpyxl.load_workbook('Input.xlsx')
ws = wb['Sheet1']
for i in range(2,172,1):
    #print(ws.cell(row=i, column=2).value)
    #print(ws.cell(row=i, column=1).value)
    links.append(ws.cell(row=i, column=2).value)
    names.append(ws.cell(row=i, column=1).value)

for link in links:
    doc = r.get(link,headers=headers)
    soup = BeautifulSoup(doc.text, "html.parser")
    for name in names:
        if names.index(name) == links.index(link):
            filename = str(name)+'.txt'
            Func = open(filename,"w+")                             # Creating a text file
            for title in soup.find_all('title'):
                Func.write(title.get_text())                       # Adding input data to the text file
                Func.write('\n')                                   # Separating title from paragraph body
            for para in soup.find_all('p'):
                Func.write(para.get_text())
            Func.close()                                       # Saving the data into the text file
            TextAnalysis(filename)

        else:
            continue

export_df = pd.read_csv("OutputData.csv", header=None)

export_df.to_csv("OutputData.csv", header=['URL_ID','URL','POSITIVE SCORE','NEGATIVE SCORE','POLARITY SCORE','SUBJECTIVITY SCORE','AVG SENTENCE LENGTH','PERCENTAGE OF COMPLEX WORDS (in %)','FOG INDEX','AVG NUMBER OF WORDS PER SENTENCE','COMPLEX WORD COUNT','WORD COUNT','SYLLABLE PER WORD','PERSONAL PRONOUNS','AVG WORD LENGTH'], index=False)
